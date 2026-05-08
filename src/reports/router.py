from fastapi import APIRouter, Depends
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.writer.excel import save_virtual_workbook
import transliterate
import re
from datetime import time
from typing import Annotated

from exceptions import InteractiveNotConductedException
from interactivities.schemas import InteractiveType
from interactivities.repository import Repository as Repository_interactive
from broadcasts.repository import Repository as Repository_broadcasts
import minios3.services as services
from config import URL_MINIO
from auth.router import get_current_active_token
from auth.schemas import TokenData

from reports.schemas import ExportGet, ExportEnum, ReturnUrl
from reports.repository import Repository

router = APIRouter(
    prefix="/api/reports",
    tags=["/api/reports"]
)


@router.post("/export")
async def get_export(
        current_token: Annotated[TokenData, Depends(get_current_active_token)],
        input_data: ExportGet
) -> ReturnUrl:
    for interactive_id in input_data.interactive_id:
        flag = await Repository.check_user_conducted_interactive(organization_id=current_token.organization_id,
                                                                 interactive_id=interactive_id.id)
        if not flag:
            raise InteractiveNotConductedException()

    bucket = "reports"

    if input_data.report_type == ExportEnum.forAnalise.value:
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"

        headers = [
            "id_интерактива", "Название интерактива", "Дата проведения",
            "Общее количество участников", "Общее количество вопросов",
            "Целевая аудитория", "Место проведения", "ФИО ведущего",
            "Способ подключения", "vk_id", "Имя", "Фамилия", "Почта", "Номер телефона",
            "Введенное имя пользователя", "Кикнут с интерактива?", "Скрыто имя на интерактиве?",
            "Количество правильных ответов", "Общее время на ответа",
            "Общее количество баллов"
        ]
        ws.append(headers)

        # Данные
        for interactive_id_data in input_data.interactive_id:
            items = await Repository.get_interactive_export_for_analise(interactive_id_data.id)
            for item in items:
                ws.append([
                    item.interactive_id,
                    item.title,
                    item.date_completed,
                    item.participant_count,
                    item.question_count,
                    item.target_audience,
                    item.location,
                    item.responsible_full_name,

                    item.provider,
                    item.vk_id,
                    item.first_name,
                    item.last_name,
                    item.email,
                    item.phone_number,

                    item.name,
                    item.is_blocked,
                    item.is_hidden,

                    item.correct_answers_count,
                    item.total_time,
                    item.total_score,
                ])

        # Возвращаем файл
        filename = "PRC_analytics_report.xlsx"
        if len(input_data.interactive_id) == 1:
            data_title_date = await Repository.get_title_and_date_for_interactive(input_data.interactive_id[0].id)
            if data_title_date:
                translit_title = smart_translit(data_title_date.title).lower().replace(' ', '_')
                translit_title = re.sub(r'[^\w_]', '', translit_title)
                filename = f"PRC_{translit_title}_{data_title_date.date_completed}.xlsx"

        file_bytes = save_virtual_workbook(wb)

        unique = await Repository_interactive.generate_unique_filename(ext="xlsx", bucket_name=bucket)

        saved_file = await services.save_image_to_minio(file=file_bytes, filename=filename, unique_filename=unique,
                                                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                        size=len(file_bytes), bucket_name=bucket)

        await Repository_broadcasts.save_image(saved_file)
        url = URL_MINIO
        return ReturnUrl(url=f"{url}{bucket}/{saved_file.unique_filename}", name=filename)

    else:
        wb = Workbook()
        wb.remove(wb.active)  # Удаляем дефолтный лист

        for interactive_id_data in input_data.interactive_id:
            data = await Repository.get_export_for_leader(interactive_id_data.id)

            # Создаем новый лист для каждого интерактива
            ws = wb.create_sheet(title=f"Интерактив {data.header.interactive_id}")

            # Настройка ширины столбцов
            for col in range(1, 14):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Стили
            correct_answer_fill = PatternFill(start_color="c1f0c8", end_color="c1f0c8", fill_type="solid")
            statistic_fill = PatternFill(start_color="f6fbc8", end_color="f6fbc8", fill_type="solid")
            statistic_time_fill = PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid")
            bold_font = Font(bold=True)
            medium_side = Side(border_style='medium', color='000000')  # Толстая граница
            thin_side = Side(border_style='thin', color='000000')  # Обычная граница

            # Подсчёт кол-во ответивших правильно на текстовый вопрос
            dict_text_true_answer = {}  # int:int | id:count

            # 1. Заголовок интерактива (A1:M7)
            interact_info = [
                f"Название интерактива: {data.header.title}",
                f"Id_интерактива: {data.header.interactive_id}",
                f"Дата проведения: {data.header.date_completed}",
                f"Общее количество участников: {len(data.body)}",
                f"Целевая аудитория: {data.header.target_audience}" if data.header.target_audience else "Целевая аудитория: не указано",
                f"Место проведения: {data.header.location}" if data.header.location else "Место проведения: не указано",
                f"ФИО ведущего: {data.header.responsible_full_name}" if data.header.responsible_full_name else "ФИО ведущего: не указано"
            ]

            for i, info in enumerate(interact_info, start=1):
                cell = ws.cell(row=i, column=1, value=info)
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=13)
                cell.font = bold_font

            # 2. Вопросы и ответы (строка 11-14)
            current_col = 14  # Начинаем с колонки J

            for question in sorted(data.header.question, key=lambda q: q.position):
                if question.type == InteractiveType.one or question.type == InteractiveType.many:
                    answer_count = len(question.answers)

                    # Заголовок вопроса (строка 11)
                    cell = ws.cell(row=11, column=current_col,
                                   value=f"Вопрос {question.position} Сложность - {question.score}")
                    cell.border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=thin_side)
                    ws.merge_cells(start_row=11, start_column=current_col, end_row=11,
                                   end_column=current_col + answer_count)

                    # Текст вопроса (строка 12)
                    cell = ws.cell(row=12, column=current_col, value=question.text)
                    cell.border = Border(top=thin_side, left=medium_side, right=medium_side, bottom=thin_side)
                    ws.merge_cells(start_row=12, start_column=current_col, end_row=12,
                                   end_column=current_col + answer_count)

                    # добавляем Время на ответ (строка 14)
                    cell = ws.cell(row=13, column=current_col + answer_count, value=f"Время на ответ")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
                    cell = ws.cell(row=14, column=current_col + answer_count, value="")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=medium_side)
                    cell = ws.cell(row=15, column=current_col + answer_count, value="")
                    cell.border = Border(top=medium_side, left=thin_side, right=medium_side, bottom=thin_side)

                    # Тексты ответов (строка 13, 14, 15)
                    for i, answer in enumerate(question.answers):
                        cell0 = ws.cell(row=13, column=current_col + i, value=f"Ответ {i + 1}")
                        cell = ws.cell(row=14, column=current_col + i, value=answer.text)
                        cell2 = ws.cell(row=15, column=current_col + i, value="")
                        if answer.is_correct:
                            cell0.fill = correct_answer_fill
                            cell.fill = correct_answer_fill
                            cell2.fill = correct_answer_fill

                        if i == 0:
                            cell0.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=medium_side)
                            cell2.border = Border(top=medium_side, left=medium_side, right=thin_side, bottom=thin_side)
                        else:
                            cell0.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                            cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                            cell2.border = Border(top=medium_side, left=thin_side, right=thin_side, bottom=thin_side)

                    current_col += answer_count + 1  # Сдвигаем на нужное количество колонок
                else:
                    answer_count = 3

                    # Заголовок вопроса (строка 11)
                    cell = ws.cell(row=11, column=current_col,
                                   value=f"Вопрос {question.position} Сложность - {question.score}")
                    cell.border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=thin_side)
                    ws.merge_cells(start_row=11, start_column=current_col, end_row=11,
                                   end_column=current_col + answer_count)

                    # Текст вопроса (строка 12)
                    cell = ws.cell(row=12, column=current_col, value=question.text)
                    cell.border = Border(top=thin_side, left=medium_side, right=medium_side, bottom=thin_side)
                    ws.merge_cells(start_row=12, start_column=current_col, end_row=12,
                                   end_column=current_col + answer_count)

                    # Варианты ответов (строка 13)
                    cell = ws.cell(row=13, column=current_col, value=f"Правильные ответы")
                    cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                    cell.fill = correct_answer_fill
                    ws.merge_cells(start_row=13, start_column=current_col, end_row=13,
                                   end_column=current_col + answer_count - 1)

                    # добавляем Время на ответ (строка 13)
                    cell = ws.cell(row=13, column=current_col + answer_count, value=f"Время на ответ")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
                    cell = ws.cell(row=14, column=current_col + answer_count, value="")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=medium_side)
                    cell = ws.cell(row=15, column=current_col + answer_count, value="")
                    cell.border = Border(top=medium_side, left=thin_side, right=medium_side, bottom=thin_side)

                    # Тексты ответов (строка 14)
                    text_correct_answer = ", ".join([answer.text for answer in question.answers])
                    cell = ws.cell(row=14, column=current_col, value=text_correct_answer)
                    cell.fill = correct_answer_fill
                    cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=medium_side)
                    ws.merge_cells(start_row=14, start_column=current_col, end_row=14,
                                   end_column=current_col + answer_count - 1)

                    cell = ws.cell(row=15, column=current_col, value="")
                    cell.border = Border(top=medium_side, left=medium_side, right=thin_side, bottom=thin_side)
                    cell.fill = correct_answer_fill
                    ws.merge_cells(start_row=15, start_column=current_col, end_row=15,
                                   end_column=current_col + answer_count - 1)

                    current_col += answer_count + 1  # Сдвигаем на нужное количество колонок

            # 2.2 Общие показатели участника (14 строчка с E по G)
            cell = ws.cell(row=14, column=11, value=f"Общие показатели участника")
            cell.border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
            ws.merge_cells(start_row=14, start_column=11, end_row=14, end_column=13)

            # 3. Заголовки таблицы участников (строка 15)
            headers = [
                "Количество участников", "Способ подключения", "vk_id", "Имя", "Фамилия", "Почта", "Номер телефона",
                "Введенное имя пользователя", "Кикнут с интерактива?", "Скрыто имя на интерактиве?",
                "Количество верных ответов", "Общее время на ответы", "Общее количество баллов"
            ]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=15, column=col, value=header)
                if header == "Количество верных ответов":
                    cell.border = Border(top=medium_side, left=medium_side, right=thin_side, bottom=thin_side)
                elif header == "Общее количество баллов":
                    cell.border = Border(top=medium_side, left=thin_side, right=medium_side, bottom=thin_side)
                else:
                    cell.border = Border(top=medium_side, left=thin_side, right=thin_side, bottom=thin_side)

            # 4. Данные участников (начиная со строки 16)

            count_participant = len(data.body)
            for i, participant in enumerate(data.body, start=1):
                row = 15 + i

                # Основная информация
                if i == count_participant:
                    cell = ws.cell(row=row, column=1, value=i)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)

                    cell = ws.cell(row=row, column=2, value=participant.provider)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)

                    cell = ws.cell(row=row, column=3, value=participant.vk_id)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=4, value=participant.first_name)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=5, value=participant.last_name)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=6, value=participant.email)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=7, value=participant.phone_number)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)

                    cell = ws.cell(row=row, column=8, value=participant.name)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=9, value=participant.is_blocked)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=10, value=participant.is_hidden)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)

                    cell = ws.cell(row=row, column=11,
                                   value=f"{participant.correct_answers_count}/{len(data.header.question)}")
                    cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=12, value=f"{participant.total_time}")
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell = ws.cell(row=row, column=13, value=f"{participant.total_score}")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=medium_side)
                else:
                    cell = ws.cell(row=row, column=1, value=i)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

                    cell = ws.cell(row=row, column=2, value=participant.provider)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

                    cell = ws.cell(row=row, column=3, value=participant.vk_id)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=4, value=participant.first_name)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=5, value=participant.last_name)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=6, value=participant.email)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=7, value=participant.phone_number)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

                    cell = ws.cell(row=row, column=8, value=participant.name)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=9, value=participant.is_blocked)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=10, value=participant.is_hidden)
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

                    cell = ws.cell(row=row, column=11,
                                   value=f"{participant.correct_answers_count}/{len(data.header.question)}")
                    cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=12, value=f"{participant.total_time}")
                    cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell = ws.cell(row=row, column=13, value=f"{participant.total_score}")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)

                # Ответы участника
                current_col = 14
                for question in sorted(data.header.question, key=lambda q: q.position):
                    if question.type == InteractiveType.one:
                        answer_count = len(question.answers)

                        # Сначала заполняем все 0
                        for j in range(answer_count):
                            cell = ws.cell(row=row, column=current_col + j, value=0)
                            if j == 0:
                                cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            else:
                                cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

                        # Затем отмечаем выбранные ответы
                        flag_is_answered = True
                        for answer in participant.answers:
                            if answer.question_id == question.id:
                                for j, a in enumerate(question.answers):
                                    if a.id == answer.answer_id:
                                        cell = ws.cell(row=row, column=current_col + j, value=1)
                                        if j == 0:
                                            cell.border = Border(top=thin_side, left=medium_side, right=thin_side,
                                                                 bottom=thin_side)
                                        else:
                                            cell.border = Border(top=thin_side, left=thin_side, right=thin_side,
                                                                 bottom=thin_side)

                                        if answer.is_correct:
                                            cell.fill = correct_answer_fill
                                        minutes, seconds = map(int, answer.time.split(':'))
                                        time_obj = time(0, minutes, seconds)  # часы, минуты, секунды
                                        cell = ws.cell(row=row, column=current_col + answer_count, value=time_obj)
                                        cell.border = Border(top=thin_side, left=thin_side, right=medium_side,
                                                             bottom=thin_side)
                                        flag_is_answered = False

                        if flag_is_answered:
                            cell = ws.cell(row=row, column=current_col + answer_count, value="")
                            cell.border = Border(top=thin_side, left=thin_side, right=medium_side,
                                                 bottom=thin_side)

                        current_col += answer_count + 1

                    elif question.type == InteractiveType.many:
                        answer_count = len(question.answers)

                        # Сначала заполняем все 0
                        for j in range(answer_count):
                            cell = ws.cell(row=row, column=current_col + j, value=0)
                            if j == 0:
                                cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            else:
                                cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

                        # Затем отмечаем выбранные ответы
                        flag_is_answered = True
                        for answer in participant.answers:
                            if answer.question_id == question.id:
                                for j, a in enumerate(question.answers):
                                    if a.id in answer.answer_id:
                                        cell = ws.cell(row=row, column=current_col + j, value=1)
                                        if j == 0:
                                            cell.border = Border(top=thin_side, left=medium_side, right=thin_side,
                                                                 bottom=thin_side)
                                        else:
                                            cell.border = Border(top=thin_side, left=thin_side, right=thin_side,
                                                                 bottom=thin_side)

                                        if answer.is_correct:
                                            cell.fill = correct_answer_fill

                                        minutes, seconds = map(int, answer.time.split(':'))
                                        time_obj = time(0, minutes, seconds)  # часы, минуты, секунды
                                        cell = ws.cell(row=row, column=current_col + answer_count, value=time_obj)
                                        cell.border = Border(top=thin_side, left=thin_side, right=medium_side,
                                                             bottom=thin_side)
                                        flag_is_answered = False

                        if flag_is_answered:
                            cell = ws.cell(row=row, column=current_col + answer_count, value="")
                            cell.border = Border(top=thin_side, left=thin_side, right=medium_side,
                                                 bottom=thin_side)

                        current_col += answer_count + 1

                    else:
                        flag_is_answered = True
                        for answer in participant.answers:
                            if answer.question_id == question.id:
                                cell = ws.cell(row=row, column=current_col, value=f"{answer.answer_id}")
                                cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                                if answer.is_correct:
                                    cell.fill = correct_answer_fill
                                    if question.id in dict_text_true_answer:
                                        dict_text_true_answer[question.id] += 1
                                    else:
                                        dict_text_true_answer[question.id] = 1
                                ws.merge_cells(start_row=row, start_column=current_col, end_row=row,
                                               end_column=current_col + 2)
                                minutes, seconds = map(int, answer.time.split(':'))
                                time_obj = time(0, minutes, seconds)  # часы, минуты, секунды
                                cell = ws.cell(row=row, column=current_col + 3, value=time_obj)
                                cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
                                flag_is_answered = False

                        if flag_is_answered:
                            cell = ws.cell(row=row, column=current_col, value="")
                            cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            ws.merge_cells(start_row=row, start_column=current_col, end_row=row,
                                           end_column=current_col + 2)
                            cell = ws.cell(row=row, column=current_col + 3, value="")
                            cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)

                        current_col += 4

            # 5. Подсчет ответивших (строка после последнего участника)
            stats_row = 15 + len(data.body) + 1 + 1
            cell = ws.cell(row=stats_row, column=11, value="Общие показатели вопроса")
            cell.font = bold_font
            cell.fill = statistic_fill
            cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
            ws.merge_cells(start_row=stats_row, start_column=11, end_row=stats_row, end_column=13)

            cell = ws.cell(row=stats_row + 1, column=11, value="Количество ответивших")
            cell.fill = statistic_fill
            cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
            ws.merge_cells(start_row=stats_row + 1, start_column=11, end_row=stats_row + 1, end_column=13)

            cell = ws.cell(row=stats_row + 2, column=11, value="Среднее время ответа на вопрос")
            cell.fill = statistic_fill
            cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
            ws.merge_cells(start_row=stats_row + 2, start_column=11, end_row=stats_row + 2, end_column=13)

            current_col = 14
            for question in sorted(data.header.question, key=lambda q: q.position):
                if question.type == InteractiveType.one or question.type == InteractiveType.many:
                    answer_count = len(question.answers)
                    first_row = 16
                    last_row = 15 + len(data.body)

                    for j, a in enumerate(question.answers):
                        col = current_col + j
                        cell = ws.cell(row=stats_row + 1, column=col,
                                       value=f"=SUM({get_column_letter(col)}{first_row}:{get_column_letter(col)}{last_row})")
                        cell2 = ws.cell(row=stats_row + 2, column=col, value="")
                        cell2.fill = statistic_fill
                        cell1 = ws.cell(row=stats_row, column=col, value="")
                        cell1.fill = statistic_fill
                        cell0 = ws.cell(row=stats_row - 1, column=col, value="")

                        if a.is_correct:
                            cell.fill = correct_answer_fill
                        else:
                            cell.fill = statistic_fill

                        if j == 0:
                            cell0.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            cell1.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            cell.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                            cell2.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=medium_side)
                        else:
                            cell0.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                            cell1.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                            cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                            cell2.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)

                    cell = ws.cell(row=stats_row - 1, column=current_col + answer_count, value="")
                    cell.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
                    cell1 = ws.cell(row=stats_row, column=current_col + answer_count, value="")
                    cell1.fill = statistic_fill
                    cell1.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)
                    cell2 = ws.cell(row=stats_row + 1, column=current_col + answer_count, value="")
                    cell2.fill = statistic_fill
                    cell2.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)

                    cell3 = ws.cell(row=stats_row + 2, column=current_col + answer_count,
                                    value=f"=AVERAGE({get_column_letter(current_col + answer_count)}{first_row}:{get_column_letter(current_col + answer_count)}{last_row})")
                    cell3.fill = statistic_time_fill
                    cell3.number_format = 'mm:ss'
                    cell3.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=medium_side)
                    current_col += answer_count + 1
                else:
                    first_row = 16
                    last_row = 15 + len(data.body)

                    # dict_text_true_answer
                    cell00 = ws.cell(row=stats_row - 1, column=current_col, value="")
                    cell01 = ws.cell(row=stats_row - 1, column=current_col + 1, value="")
                    cell02 = ws.cell(row=stats_row - 1, column=current_col + 2, value="")
                    cell03 = ws.cell(row=stats_row - 1, column=current_col + 3, value="")
                    cell00.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                    cell01.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell02.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell03.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)

                    cell10 = ws.cell(row=stats_row, column=current_col, value="")
                    cell11 = ws.cell(row=stats_row, column=current_col + 1, value="")
                    cell12 = ws.cell(row=stats_row, column=current_col + 2, value="")
                    cell13 = ws.cell(row=stats_row, column=current_col + 3, value="")
                    cell10.fill = statistic_fill
                    cell11.fill = statistic_fill
                    cell12.fill = statistic_fill
                    cell13.fill = statistic_fill
                    cell10.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                    cell11.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell12.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell13.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)

                    cell20 = ws.cell(row=stats_row + 1, column=current_col, value="")
                    if question.id in dict_text_true_answer:
                        cell21 = ws.cell(row=stats_row + 1, column=current_col + 1,
                                         value=dict_text_true_answer[question.id])
                    else:
                        cell21 = ws.cell(row=stats_row + 1, column=current_col + 1, value=0)
                    cell22 = ws.cell(row=stats_row + 1, column=current_col + 2, value="")
                    cell23 = ws.cell(row=stats_row + 1, column=current_col + 3, value="")
                    cell20.fill = statistic_fill
                    cell21.fill = correct_answer_fill
                    cell22.fill = statistic_fill
                    cell23.fill = statistic_fill
                    cell20.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=thin_side)
                    cell21.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell22.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    cell23.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=thin_side)

                    cell30 = ws.cell(row=stats_row + 2, column=current_col, value="")
                    cell31 = ws.cell(row=stats_row + 2, column=current_col + 1, value="")
                    cell32 = ws.cell(row=stats_row + 2, column=current_col + 2, value="")
                    cell30.fill = statistic_fill
                    cell31.fill = statistic_fill
                    cell32.fill = statistic_fill
                    cell30.border = Border(top=thin_side, left=medium_side, right=thin_side, bottom=medium_side)
                    cell31.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell32.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
                    cell33 = ws.cell(row=stats_row + 2, column=current_col + 3,
                                     value=f"=AVERAGE({get_column_letter(current_col + 3)}{first_row}:{get_column_letter(current_col + 3)}{last_row})")
                    cell33.fill = statistic_time_fill
                    cell33.border = Border(top=thin_side, left=thin_side, right=medium_side, bottom=medium_side)
                    cell33.number_format = 'mm:ss'

                    current_col += 4

        filename = "LDR_leader_report.xlsx"
        if len(input_data.interactive_id) == 1:
            data_title_date = await Repository.get_title_and_date_for_interactive(input_data.interactive_id[0].id)
            if data_title_date:
                translit_title = smart_translit(data_title_date.title).lower().replace(' ', '_')
                translit_title = re.sub(r'[^\w_]', '', translit_title)
                filename = f"LDR_{translit_title}_{data_title_date.date_completed}.xlsx"

        file_bytes = save_virtual_workbook(wb)

        unique = await Repository_interactive.generate_unique_filename(ext="xlsx", bucket_name=bucket)

        saved_file = await services.save_image_to_minio(file=file_bytes, filename=filename, unique_filename=unique,
                                                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                        size=len(file_bytes), bucket_name=bucket)

        await Repository_broadcasts.save_image(saved_file)
        url = URL_MINIO
        return ReturnUrl(url=f"{url}{bucket}/{saved_file.unique_filename}", name=filename)


def smart_translit(text):
    # Разделяем текст на слова и символы
    words = re.findall(r'([а-яА-ЯёЁ]+|\w+|[^\w\s]+|\s+)', text)
    result = []

    for word in words:
        # Если слово содержит кириллицу — транслитерируем
        if re.search(r'[а-яА-ЯёЁ]', word):
            try:
                # Указываем язык явно (русский) и включаем строгий режим
                translit_word = transliterate.translit(word, 'ru', reversed=True)
                # Заменяем мягкий/твёрдый знаки на апостроф или удаляем
                translit_word = translit_word.replace("'", "").replace('"', '')
                result.append(translit_word)
            except Exception as e:
                print(f"Transliteration error for '{word}': {e}")
                result.append(word)  # Если ошибка — оставляем как есть
        else:
            result.append(word)  # Английские слова и символы оставляем

    return ''.join(result)
